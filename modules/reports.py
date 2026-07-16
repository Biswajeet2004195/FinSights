import os
import tkinter as tk
from tkinter import messagebox
import csv
from datetime import datetime, timedelta
from collections import defaultdict
from config import *
from base_dashboard import BaseDashboard

class ReportMixin(BaseDashboard):
    def show_reports(self):
        self._clear(); self._set_nav("Reports"); self._set_title("Reports")
        pg = self._scrollable(self._cf)
        self._sec(pg, "📋 Financial Reports", "Monthly analysis and data export")

        trans = _ld("transactions"); cm = curr_m()
        dc = GLOBAL_STATE["display_currency"]
        mi = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "income"  and r["date"].startswith(cm))
        me = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "expense" and r["date"].startswith(cm))
        ms = mi - me; sr = ms / mi * 100 if mi else 0

        srow = ctk.CTkFrame(pg, fg_color=BG, corner_radius=10); srow.pack(fill="x", padx=20, pady=(4, 12))
        self._kpi(srow, "Income",   fmt_disp(mi), "This month",     GR, "💰").pack(side="left", ipadx=8, ipady=4, padx=(0, 10), expand=True, fill="both")
        self._kpi(srow, "Expenses", fmt_disp(me), "This month",     RE, "💸").pack(side="left", ipadx=8, ipady=4, padx=(0, 10), expand=True, fill="both")
        self._kpi(srow, "Savings",  fmt_disp(ms), f"Rate: {sr:.0f}%", CY, "💎").pack(side="left", ipadx=8, ipady=4, expand=True, fill="both")

        # Bar chart card
        ccard = ctk.CTkFrame(pg, fg_color=BD, corner_radius=10); ccard.pack(fill="x", padx=20, pady=(0, 12))
        ccardi = ctk.CTkFrame(ccard, fg_color=CB, corner_radius=10); ccardi.pack(padx=1, pady=1, fill="x")
        ctk.CTkFrame(ccardi, fg_color=AC, height=3, corner_radius=10).pack(fill="x")
        tk.Label(ccardi, text="📊  Income vs Expenses — Last 6 Months",
                 font=("Segoe UI", 11, "bold"), bg=CB, fg=TP).pack(anchor="w", padx=14, pady=(10, 4))

        months_data = []
        for mo in range(5, -1, -1):
            d = datetime.now() - timedelta(days=30 * mo)
            ym = d.strftime("%Y-%m"); lbl = d.strftime("%b")
            inc = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "income"  and r["date"].startswith(ym))
            exp = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "expense" and r["date"].startswith(ym))
            months_data.append((lbl, inc, exp))

        bc = tk.Canvas(ccardi, height=240, bg=CB, bd=0, highlightthickness=0)
        bc.pack(fill="x", padx=14, pady=(4, 8))
        self.root.after(80, lambda: self._draw_bar_chart(bc, months_data))
        bc.bind("<Configure>", lambda e, c=bc, d=months_data: self._draw_bar_chart(c, d))

        leg = ctk.CTkFrame(ccardi, fg_color=CB, corner_radius=10); leg.pack(padx=14, pady=(0, 12))
        ctk.CTkFrame(leg, fg_color=GR, width=18, height=10, corner_radius=10).pack(side="left", pady=4)
        tk.Label(leg, text=" Income", font=("Segoe UI Light", 9), bg=CB, fg=TS).pack(side="left")
        ctk.CTkFrame(leg, fg_color=RE, width=18, height=10, corner_radius=10).pack(side="left", padx=(20, 0), pady=4)
        tk.Label(leg, text=" Expenses", font=("Segoe UI Light", 9), bg=CB, fg=TS).pack(side="left")

        # Category table
        tk.Label(pg, text="📊  Category Breakdown — Current Month",
                 font=("Segoe UI Semibold", 12), bg=BG, fg=TP).pack(anchor="w", padx=20, pady=(8, 4))
        cols = ["Category", "Transactions", "Total Spent", "Budget", "% of Budget"]
        tf, tv = self._tv(pg, cols, [200, 120, 140, 140, 120], height=10)
        tf.pack(fill="x", padx=20, pady=(0, 12))
        
        budgets = _ldd("budgets")
        cats = list(budgets.keys())
        cat_data = defaultdict(lambda: {"count": 0, "total": 0.0})
        for r in trans:
            if r["type"] == "expense" and r["date"].startswith(cm):
                cat_data[r["category"]]["count"]  += 1
                cat_data[r["category"]]["total"]  += convert_currency(r["amount"], r.get("currency", "INR"), dc)
        rows = []
        for cat in cats:
            d  = cat_data[cat]
            b_val = budgets.get(cat, {"amount": 0.0, "currency": "INR"})
            if isinstance(b_val, (int, float)): b_val = {"amount": float(b_val), "currency": "INR"}
            bgt = convert_currency(b_val["amount"], b_val["currency"], dc)
            pct = d["total"] / bgt * 100 if bgt else 0
            rows.append((cat, cat, str(d["count"]), fmt_disp(d["total"]),
                         fmt_disp(bgt), f"{pct:.0f}%"))
        self._tv_fill(tv, rows)

        # Advanced Export Configuration
        tk.Label(pg, text="⚙️  Export Configuration", font=("Segoe UI Semibold", 12), bg=BG, fg=TP).pack(anchor="w", padx=20, pady=(12, 4))
        
        opt_f = ctk.CTkFrame(pg, fg_color=BD, corner_radius=10); opt_f.pack(fill="x", padx=20, pady=(0, 12))
        opt_fi = ctk.CTkFrame(opt_f, fg_color=CB, corner_radius=10); opt_fi.pack(padx=1, pady=1, fill="x")
        
        # Date Range
        date_f = tk.Frame(opt_fi, bg=CB)
        date_f.pack(fill="x", padx=14, pady=(14, 4))
        
        tk.Label(date_f, text="Start Date (YYYY-MM-DD):", font=("Segoe UI", 10), bg=CB, fg=TS).pack(side="left")
        self.start_dt_var = tk.StringVar(value=f"{datetime.now().strftime('%Y-%m')}-01")
        tk.Entry(date_f, textvariable=self.start_dt_var, bg=EN, fg=TP, font=("Segoe UI", 10), width=12, insertbackground=CY, bd=0).pack(side="left", padx=10, ipady=4)
        
        tk.Label(date_f, text="End Date:", font=("Segoe UI", 10), bg=CB, fg=TS).pack(side="left", padx=(10, 0))
        self.end_dt_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        tk.Entry(date_f, textvariable=self.end_dt_var, bg=EN, fg=TP, font=("Segoe UI", 10), width=12, insertbackground=CY, bd=0).pack(side="left", padx=10, ipady=4)
        
        # Export Selection
        chk_f = tk.Frame(opt_fi, bg=CB)
        chk_f.pack(fill="x", padx=14, pady=(4, 14))
        
        self.inc_exp_var = tk.BooleanVar(value=True)
        self.inc_bgt_var = tk.BooleanVar(value=True)
        self.inc_inv_var = tk.BooleanVar(value=True)
        self.inc_gol_var = tk.BooleanVar(value=True)
        
        ctk.CTkCheckBox(chk_f, text="Transactions", variable=self.inc_exp_var, text_color=TP, fg_color=GR, hover_color=AC).pack(side="left", padx=(0, 10))
        ctk.CTkCheckBox(chk_f, text="Budgets", variable=self.inc_bgt_var, text_color=TP, fg_color=GR, hover_color=AC).pack(side="left", padx=10)
        ctk.CTkCheckBox(chk_f, text="Investments", variable=self.inc_inv_var, text_color=TP, fg_color=GR, hover_color=AC).pack(side="left", padx=10)
        ctk.CTkCheckBox(chk_f, text="Goals", variable=self.inc_gol_var, text_color=TP, fg_color=GR, hover_color=AC).pack(side="left", padx=10)
        
        # Export Actions
        ew = ctk.CTkFrame(pg, fg_color=BG, corner_radius=10); ew.pack(fill="x", padx=20, pady=(4, 20))
        self._tb_btn(ew, "⬇  Export PDF Report", self._export_pdf_report, GR)
        self._tb_btn(ew, "📊  Export Excel Report", self._export_excel_report, CY)
        self._tb_btn(ew, "📁  Export CSV", self._export_all, CB2)

    def _get_filtered_data(self):
        try:
            sd = datetime.strptime(self.start_dt_var.get(), "%Y-%m-%d")
            ed = datetime.strptime(self.end_dt_var.get(), "%Y-%m-%d")
            ed = ed.replace(hour=23, minute=59, second=59)
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD.")
            return None
            
        trans = _ld("transactions")
        filtered_trans = []
        for r in trans:
            try:
                td = datetime.strptime(r["date"], "%Y-%m-%d")
                if sd <= td <= ed:
                    filtered_trans.append(r)
            except ValueError:
                pass
                
        return {
            "trans": filtered_trans,
            "budgets": _ldd("budgets") if self.inc_bgt_var.get() else {},
            "invs": _ld("investments") if self.inc_inv_var.get() else [],
            "goals": _ld("goals") if self.inc_gol_var.get() else [],
            "start": sd.strftime("%Y-%m-%d"),
            "end": ed.strftime("%Y-%m-%d")
        }

    def _get_exports_path(self):
        export_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
        os.makedirs(export_dir, exist_ok=True)
        return export_dir

    def _gen_charts(self, trans):
        import io
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        
        charts = {}
        dc = GLOBAL_STATE["display_currency"]
        inc = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "income")
        exp = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in trans if r["type"] == "expense")
        
        if inc > 0 or exp > 0:
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.bar(['Income', 'Expenses'], [inc, exp], color=['#10b981', '#ff453a'])
            ax.set_title('Income vs Expenses')
            buf = io.BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight')
            buf.seek(0)
            charts['inc_exp'] = buf
            plt.close(fig)
            
        spent = defaultdict(float)
        for r in trans:
            if r["type"] == "expense":
                spent[r["category"]] += convert_currency(r["amount"], r.get("currency", "INR"), dc)
                
        if spent:
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.pie(spent.values(), labels=spent.keys(), autopct='%1.1f%%', startangle=90)
            ax.set_title('Expense Breakdown')
            buf = io.BytesIO()
            fig.savefig(buf, format='png', bbox_inches='tight')
            buf.seek(0)
            charts['pie'] = buf
            plt.close(fig)
            
        return charts

    def _export_all(self):
        data = self._get_filtered_data()
        if not data: return
        
        fn = os.path.join(self._get_exports_path(), f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        
        with open(fn, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([f"Finsights Report ({data['start']} to {data['end']})"])
            w.writerow([])
            dc = GLOBAL_STATE["display_currency"]
            
            if self.inc_exp_var.get():
                w.writerow(["--- TRANSACTIONS ---"])
                w.writerow(["Date", "Type", "Description", "Category", f"Amount ({dc})", "Notes"])
                for r in sorted(data["trans"], key=lambda x: x["date"], reverse=True):
                    val = convert_currency(r["amount"], r.get("currency", "INR"), dc)
                    w.writerow([r["date"], r["type"], r["desc"], r["category"], val, r.get("notes", "")])
                w.writerow([])
                
            if self.inc_bgt_var.get():
                w.writerow(["--- BUDGETS ---"])
                w.writerow(["Category", f"Budget Amount ({dc})"])
                for k, v in data["budgets"].items():
                    if isinstance(v, (int, float)): v = {"amount": float(v), "currency": "INR"}
                    val = convert_currency(v["amount"], v["currency"], dc)
                    w.writerow([k, val])
                w.writerow([])
                
            if self.inc_inv_var.get():
                w.writerow(["--- INVESTMENTS ---"])
                w.writerow(["Name", "Symbol", "Type", "Qty", f"Buy Price ({dc})", f"Current Price ({dc})"])
                for i in data["invs"]:
                    b_val = convert_currency(i["buy_price"], i.get("currency", "INR"), dc)
                    c_val = convert_currency(i["current_price"], i.get("currency", "INR"), dc)
                    w.writerow([i["name"], i["symbol"], i["type"], i["qty"], b_val, c_val])
                w.writerow([])
                
            if self.inc_gol_var.get():
                w.writerow(["--- GOALS ---"])
                w.writerow(["Name", f"Target ({dc})", f"Saved ({dc})", "Deadline"])
                for g in data["goals"]:
                    t_val = convert_currency(float(g["target"]), g.get("currency", "INR"), dc)
                    s_val = convert_currency(float(g["saved"]), g.get("currency", "INR"), dc)
                    w.writerow([g["name"], t_val, s_val, g.get("deadline", "")])
                    
        messagebox.showinfo("✅ Exported", f"CSV report saved to:\n{fn}")

    def _export_pdf_report(self):
        data = self._get_filtered_data()
        if not data: return
        
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
        except ImportError:
            return messagebox.showerror("Error", "ReportLab is not installed.")

        fn = os.path.join(self._get_exports_path(), f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        doc = SimpleDocTemplate(fn, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()
        
        primary_color = colors.HexColor("#0a84ff")
        dark_neutral = colors.HexColor("#161617")
        light_neutral = colors.HexColor("#f4f4f5")
        
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=22, textColor=primary_color, spaceAfter=6)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor("#8e8e93"), spaceAfter=15)
        heading_style = ParagraphStyle('HeadingStyle', parent=styles['Heading2'], fontSize=13, textColor=dark_neutral, spaceBefore=14, spaceAfter=8)
        body_header_style = ParagraphStyle('BodyHeaderStyle', parent=styles['Normal'], fontSize=9, textColor=colors.white, fontName="Helvetica-Bold")
        
        story.append(Paragraph("Finsights — Financial Report", title_style))
        story.append(Paragraph(f"Period: {data['start']} to {data['end']}", subtitle_style))
        
        # Summary
        dc = GLOBAL_STATE["display_currency"]
        if self.inc_exp_var.get():
            mi = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in data["trans"] if r["type"] == "income")
            me = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in data["trans"] if r["type"] == "expense")
            ms = mi - me
            sr = ms / mi * 100 if mi else 0
            
            summary_data = [
                [Paragraph("Total Income", body_header_style), Paragraph("Total Expenses", body_header_style), 
                 Paragraph("Net Savings", body_header_style), Paragraph("Savings Rate", body_header_style)],
                [fmt_disp(mi), fmt_disp(me), fmt_disp(ms), f"{sr:.1f}%"]
            ]
            summary_table = Table(summary_data, colWidths=[135]*4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('BACKGROUND', (0,1), (-1,1), light_neutral),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e4e4e7")),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(Paragraph("Financial Summary", heading_style))
            story.append(summary_table)
            story.append(Spacer(1, 12))
            
            # Add Charts
            charts = self._gen_charts(data["trans"])
            chart_data = []
            row = []
            if 'inc_exp' in charts:
                row.append(RLImage(charts['inc_exp'], width=240, height=180))
            if 'pie' in charts:
                row.append(RLImage(charts['pie'], width=240, height=180))
            if row:
                story.append(Table([row], colWidths=[270]*len(row)))
                story.append(Spacer(1, 12))
                
            # Transactions
            story.append(Paragraph("Transactions", heading_style))
            tx_data = [[Paragraph("Date", body_header_style), Paragraph("Type", body_header_style), 
                        Paragraph("Description", body_header_style), Paragraph("Category", body_header_style), 
                        Paragraph("Amount", body_header_style)]]
            for r in sorted(data["trans"], key=lambda x: x["date"], reverse=True)[:50]:
                tx_data.append([r["date"], r["type"].capitalize(), r["desc"][:25], r["category"], fmt_amt(r['amount'], r.get('currency', 'INR'))])
            
            tx_table = Table(tx_data, colWidths=[80, 70, 160, 110, 120])
            tx_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), dark_neutral),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e4e4e7")),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_neutral]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(tx_table)
            
        if self.inc_bgt_var.get() and data["budgets"]:
            story.append(Paragraph("Budgets", heading_style))
            bd_data = [[Paragraph("Category", body_header_style), Paragraph("Budget", body_header_style)]]
            for k, v in data["budgets"].items():
                if isinstance(v, (int, float)): v = {"amount": float(v), "currency": "INR"}
                bd_data.append([k, fmt_amt(v["amount"], v["currency"])])
            bd_table = Table(bd_data, colWidths=[270, 270])
            bd_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), dark_neutral),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e4e4e7")),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_neutral]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(bd_table)
            
        if self.inc_inv_var.get() and data["invs"]:
            story.append(Paragraph("Investments", heading_style))
            inv_data = [[Paragraph("Name", body_header_style), Paragraph("Type", body_header_style), 
                         Paragraph("Qty", body_header_style), Paragraph("Buy Price", body_header_style), 
                         Paragraph("Current Value", body_header_style)]]
            for i in data["invs"]:
                curr = i.get("currency", "INR")
                inv_data.append([i["name"], i["type"], str(i["qty"]), fmt_amt(i["buy_price"], curr), fmt_amt(i["current_price"], curr)])
            inv_table = Table(inv_data, colWidths=[120, 100, 70, 120, 130])
            inv_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), dark_neutral),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e4e4e7")),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_neutral]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(inv_table)
            
        if self.inc_gol_var.get() and data["goals"]:
            story.append(Paragraph("Goals", heading_style))
            gol_data = [[Paragraph("Name", body_header_style), Paragraph("Target", body_header_style), 
                         Paragraph("Saved", body_header_style), Paragraph("Deadline", body_header_style)]]
            for g in data["goals"]:
                curr = g.get("currency", "INR")
                gol_data.append([g["name"], fmt_amt(float(g["target"]), curr), fmt_amt(float(g["saved"]), curr), g.get("deadline", "")])
            gol_table = Table(gol_data, colWidths=[160, 120, 120, 140])
            gol_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), dark_neutral),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e4e4e7")),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_neutral]),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(gol_table)
            
        doc.build(story)
        messagebox.showinfo("✅ Exported", f"PDF report saved to:\n{fn}")

    def _export_excel_report(self):
        data = self._get_filtered_data()
        if not data: return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            return messagebox.showerror("Error", "openpyxl is not installed.")
            
        fn = os.path.join(self._get_exports_path(), f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        title_font = Font(name="Segoe UI", size=16, bold=True, color="0A84FF")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="161617", end_color="161617", fill_type="solid")
        thin_side = Side(border_style="thin", color="E4E4E7")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws["A1"] = f"Finsights Report ({data['start']} to {data['end']})"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 30
        
        dc = GLOBAL_STATE["display_currency"]
        
        if self.inc_exp_var.get():
            mi = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in data["trans"] if r["type"] == "income")
            me = sum(convert_currency(r["amount"], r.get("currency", "INR"), dc) for r in data["trans"] if r["type"] == "expense")
            ms = mi - me
            
            ws["A3"] = "Total Income"; ws["B3"] = mi
            ws["A4"] = "Total Expenses"; ws["B4"] = me
            ws["A5"] = "Net Savings"; ws["B5"] = ms
            for r in range(3, 6):
                ws.cell(row=r, column=1).font = Font(bold=True)
                ws.cell(row=r, column=2).number_format = '0.00'
                
            charts = self._gen_charts(data["trans"])
            if 'inc_exp' in charts:
                img = XLImage(charts['inc_exp'])
                ws.add_image(img, 'D3')
            if 'pie' in charts:
                img = XLImage(charts['pie'])
                ws.add_image(img, 'D18')
                
            ws_tx = wb.create_sheet(title="Transactions")
            headers = ["Date", "Type", "Description", "Category", "Amount", "Notes"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws_tx.cell(row=1, column=col_idx, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.border = border_all
                
            for row_idx, r in enumerate(sorted(data["trans"], key=lambda x: x["date"], reverse=True), 2):
                ws_tx.cell(row=row_idx, column=1, value=r["date"])
                ws_tx.cell(row=row_idx, column=2, value=r["type"])
                ws_tx.cell(row=row_idx, column=3, value=r["desc"])
                ws_tx.cell(row=row_idx, column=4, value=r["category"])
                ws_tx.cell(row=row_idx, column=5, value=convert_currency(r["amount"], r.get("currency", "INR"), dc)).number_format = '0.00'
                ws_tx.cell(row=row_idx, column=6, value=r.get("notes", ""))
                
        if self.inc_bgt_var.get() and data["budgets"]:
            ws_bd = wb.create_sheet(title="Budgets")
            headers = ["Category", "Budget Limit"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws_bd.cell(row=1, column=col_idx, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.border = border_all
            for row_idx, (k, v) in enumerate(data["budgets"].items(), 2):
                if isinstance(v, (int, float)): v = {"amount": float(v), "currency": "INR"}
                ws_bd.cell(row=row_idx, column=1, value=k)
                ws_bd.cell(row=row_idx, column=2, value=convert_currency(v["amount"], v["currency"], dc)).number_format = '0.00'
                
        if self.inc_inv_var.get() and data["invs"]:
            ws_inv = wb.create_sheet(title="Investments")
            headers = ["Name", "Symbol", "Type", "Qty", "Buy Price", "Current Price"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws_inv.cell(row=1, column=col_idx, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.border = border_all
            for row_idx, i in enumerate(data["invs"], 2):
                curr = i.get("currency", "INR")
                ws_inv.cell(row=row_idx, column=1, value=i["name"])
                ws_inv.cell(row=row_idx, column=2, value=i["symbol"])
                ws_inv.cell(row=row_idx, column=3, value=i["type"])
                ws_inv.cell(row=row_idx, column=4, value=i["qty"])
                ws_inv.cell(row=row_idx, column=5, value=convert_currency(i["buy_price"], curr, dc)).number_format = '0.00'
                ws_inv.cell(row=row_idx, column=6, value=convert_currency(i["current_price"], curr, dc)).number_format = '0.00'
                
        if self.inc_gol_var.get() and data["goals"]:
            ws_gol = wb.create_sheet(title="Goals")
            headers = ["Name", "Target", "Saved", "Deadline"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws_gol.cell(row=1, column=col_idx, value=h)
                cell.font = header_font; cell.fill = header_fill; cell.border = border_all
            for row_idx, g in enumerate(data["goals"], 2):
                curr = g.get("currency", "INR")
                ws_gol.cell(row=row_idx, column=1, value=g["name"])
                ws_gol.cell(row=row_idx, column=2, value=convert_currency(float(g["target"]), curr, dc)).number_format = '0.00'
                ws_gol.cell(row=row_idx, column=3, value=convert_currency(float(g["saved"]), curr, dc)).number_format = '0.00'
                ws_gol.cell(row=row_idx, column=4, value=g.get("deadline", ""))
                
        for s in wb.sheetnames:
            worksheet = wb[s]
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(fn)
        messagebox.showinfo("✅ Exported", f"Excel report saved to:\n{fn}")
